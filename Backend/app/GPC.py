#!/usr/bin/env python3
import sys, json, logging
from pathlib import Path
import pandas as pd
from zipfile import ZipFile
import openpyxl
from openpyxl import Workbook
import win32com.client
import pyodbc
from datetime import datetime
import os
from APIData import strategy_direct_json
from dotenv import load_dotenv
load_dotenv()
# ─── CONFIG ────────────────────────────────────────────────────────────────
BASE_DIR    = Path(os.getenv("GPC_BASE_PATH"))
INPUT_DIR   = BASE_DIR / "Data"
OUTPUT_DIR  = BASE_DIR / "Output"
ACCDB       = BASE_DIR / "IFS.accdb"
TEMPLATE    = INPUT_DIR / "Import_Florecompc2_template.json"
DATE_STR = datetime.now().strftime("%Y%m%d")
# which row to start inserting CSV rows into
START_ROW   = 4  
START_COL   = 1  

# sheet → filename-pattern
SHEET_PATTERNS = {
    "Benaming":        "CN??????.txt",
    "Benamingstype":   "CM??????.txt",
    "Cultivar":        "CC??????.txt",
    "Geslacht":        "CG??????.txt",
    "Gewas":           "CT??????.txt",
    "Groep":           "CO??????.txt",
    "Soort":           "CS??????.txt",
    "Kenmerkgroep":    "CU??????.txt",
    "Kenmerktype":     "CE??????.txt",
    "Kenmerkwaarde":   "CV??????.txt",
    "Product":         "CP??????.txt",
    "Productkenmerk":  "CF??????.txt",
    "Regl. Kenmerktype":"CY??????.txt",
    "Toepassing":      "CA??????.txt",
    "Voorschrift type":"CR??????.txt",
}

QUERY_NAMES = [
    "01Plant_Genus_Species_Product_GS_RegiD",
    "02Plant_Genus_Species_Product_GS",
    "03Plant_Genus_Species_Product_G_RegiD",
    "04Plant_Genus_Species_Product_G",
    "05Plant_Genus_Species_Product_product_name_GS",
    "06Plant_Genus_Species_Product_product_name_G",
    "07Plant_Genus_Species_Product_Specific_Bricks",
    "08Plant_Genus_Species_Product_Others",
]

CODE_LIST_UPDATES = {
    'COLOR':'20','NAME_GPC':'21','NAME_TYPE':'22',
    'SEGMENT':'30','FAMILY':'31','CLASS':'32','BRICK':'33',
    'ATTRIBUTE_TYPE':'34','ATTRIBUTE_VALUE':'35',
    'SEGMENT_FAMILY_CLASS_BRICK':'36',
    'BRICK_ATTRIBUTE_TYPE_ATTRIBUTE_VALUE':'37',
    'PRODUCT_GPC':'39'
}

TABLES_TO_EXPORT = list(CODE_LIST_UPDATES.keys())

# ─── LOGGING ──────────────────────────────────────────────────────────────
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s %(levelname)s: %(message)s",
    handlers=[logging.StreamHandler(sys.stdout)]
)
logger = logging.getLogger()

def load_header_template():
    if not TEMPLATE.exists():
        logger.error(f"Template niet gevonden: {TEMPLATE}")
        sys.exit(1)
    return json.loads(TEMPLATE.read_text(encoding="utf-8"))

def excel_import_python():
    """ Build a new .xlsx from headers in JSON + all .txt data. """
    logger.info("=== Excel import (from JSON-template) ===")
    header_tpl = load_header_template()
    wb = Workbook()
    wb.remove(wb.active)

    for sheet, pattern in SHEET_PATTERNS.items():
        ws = wb.create_sheet(sheet)
        # 1) write header rows 1–3
        for r, row in enumerate(header_tpl.get(sheet, []), start=1):
            for c, v in enumerate(row, start=1):
                if v is not None:
                    ws.cell(row=r, column=c, value=v)

        # 2) import each matching .txt
        row_cursor = START_ROW
        files = sorted(INPUT_DIR.glob(pattern))
        if not files:
            logger.warning(f"Geen bestanden voor {sheet} ({pattern})")
        for f in files:
            try:
                df = pd.read_csv(f, sep=';', header=None, encoding='cp1252', low_memory=False)
            except Exception as e:
                logger.error(f"Leesfout {f.name}: {e}")
                continue

            for i, row in df.iterrows():
                for j, val in enumerate(row, start=START_COL):
                    ws.cell(row=row_cursor + i, column=j, value=val)

            # write file-name & date in E1/E2
            fn = f.name
            ws.cell(row=1, column=5, value=fn)
            dp = fn[:8]
            ws.cell(row=2, column=5, value=f"'{dp[2:4]}-{dp[4:6]}-20{dp[6:8]}")
            row_cursor += len(df)

        logger.info(f"  → Sheet '{sheet}' gevuld")

    out = INPUT_DIR / f"Import_Florecompc2_{Path().cwd().name}.xlsx"
    try:
        wb.save(out)
        logger.info(f"Nieuw workbook: {out}")
    except Exception as e:
        logger.error(f"Opslaan mislukt: {e}")
        sys.exit(1)

    return out


def run_access_queries():
    """ Open IFS.accdb, run code_list updates, action-queries, then export. """
    logger.info("=== Access queries & exports ===")
    if not ACCDB.exists():
        logger.error(f"IFS.accdb niet gevonden onder {ACCDB}")
        sys.exit(1)

    # --- DAO connectie ---
    try:
        dao = win32com.client.Dispatch("DAO.DBEngine.120")
        ws_dao = dao.Workspaces(0)
        db = ws_dao.OpenDatabase(str(ACCDB))
    except Exception as e:
        logger.error(f"DAO connectie mislukt: {e}")
        

    # 1) code_list_id updates
    ws_dao.BeginTrans()
    try:
        for tbl, code in CODE_LIST_UPDATES.items():
            sql = f"UPDATE [{tbl}] SET code_list_id='{code}';"
            db.Execute(sql)
            logger.info(f" code_list_id {tbl}→{code}")
        ws_dao.CommitTrans()
    except Exception as e:
        ws_dao.Rollback()
        logger.error(f"Rollback code_list updates: {e}")

    # 2) action-queries
    for q in QUERY_NAMES:
        try:
            db.QueryDefs(q).Execute()
            logger.info(f"Query {q} uitgevoerd")
        except Exception as e:
            logger.error(f"Fout in action-query {q}: {e}")

    # 3) export via ODBC
    OUTPUT_DIR.mkdir(exist_ok=True)
    conn_str = f"DRIVER={{Microsoft Access Driver (*.mdb, *.accdb)}};DBQ={ACCDB};"
    try:
        conn = pyodbc.connect(conn_str)
    except Exception as e:
        logger.error(f"ODBC connect mislukt: {e}")
        db.Close()
        sys.exit(1)
    with ZipFile(BASE_DIR /'Zipped file.zip', 'w') as zip_object:
        for tbl, code in CODE_LIST_UPDATES.items():
            try:
                df = pd.read_sql(f"SELECT * FROM [{tbl}]", conn)
                fn = OUTPUT_DIR / f"C{code}_{DATE_STR}.txt"
                df.to_csv(fn, sep=';', index=False)
                logger.info(f"  → Export {tbl} → {fn.name}")
                
                zip_object.write(fn)
            except Exception as e:
                logger.error(f"Export {tbl} mislukt: {e}")

    conn.close()
    db.Close()
    logger.info("Access gesloten")

if __name__ == "__main__":
    strategy_direct_json()
    # 1) build the new XLSX from JSON + .txt  
    excel_import_python()
    # 2) open Access, run queries & export
    run_access_queries()
    logger.info("🎉 Klaar!")  
