import json
from pathlib import Path
from openpyxl import load_workbook
from openpyxl import Workbook
import json
from pathlib import Path
def extract_headers(xlsm_path, rows_to_extract=(1, 2, 3), output_path=None):
    wb = load_workbook(xlsm_path, keep_vba=True, data_only=True)
    header_template = {}
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        # Extract specified rows
        extracted = []
        for r in rows_to_extract:
            row_vals = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
            extracted.append(row_vals)
        header_template[sheet_name] = extracted
    
    # Print to console
    print(json.dumps(header_template, indent=2, ensure_ascii=False))
    
    # Save to file if requested
    if output_path:
        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump(header_template, f, indent=2, ensure_ascii=False)
        print(f"Header template saved to {output_path}")


#extract_headers(xlsm, output_path=output_json)

def load_header_template(mut_dir: Path) -> dict:
    tpl_path = mut_dir/ 'header_template.json'
    with open(tpl_path, encoding='utf-8') as f:
        return json.load(f)

def create_import_workbook(mut_dir: Path, header_tpl: dict):
    inp = mut_dir / 'Input'
    wb = Workbook()
    wb.remove(wb.active)

    for sheet_name, header_rows in header_tpl.items():
        ws = wb.create_sheet(title=sheet_name)
        # Schrijf rijen 1–3 exact zoals in de template
        for i, row in enumerate(header_rows, start=1):
            for j, val in enumerate(row, start=1):
                # soms staan er nulls in de JSON; skip die
                if val is not None:
                    ws.cell(row=i, column=j, value=val)

    # Sla op
    dst = inp / f"Import_Florecompc2_{mut_dir.name}.xlsx"
    wb.save(dst)
    return dst
# Usage
mut_dir = Path(r"C:\Users\c.elkhattabi\Desktop\Test\20250423")
xlsm = mut_dir / "Input" / "Import Florecompc2.xlsm"
output_json = mut_dir / "header_template.json"
create_import_workbook(mut_dir,load_header_template(mut_dir))